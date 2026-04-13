# -*- coding: utf-8 -*-
"""
Generador de Diagramas de Ishikawa (Espina de Pescado) combinado con 5 Porqués usando OpenAI API.

Este módulo proporciona la clase IshikawaGenerator que permite generar análisis
de causa raíz utilizando la metodología de Ishikawa (6M) combinada con 5 Porqués
para profundizar en cada causa identificada.

Clases
------
IshikawaGenerator
    Generador principal de análisis de causa raíz con Ishikawa + 5 Porqués

Categorías de Ishikawa (6M)
----------------------------
1. Mano de Obra (Personal)
2. Métodos (Procedimientos)
3. Máquinas (Equipos)
4. Materiales (Insumos)
5. Medición (Control)
6. Medio Ambiente (Entorno)

Ejemplos
--------
>>> from ishikawa_generator import IshikawaGenerator
>>> 
>>> # Crear generador
>>> generator = IshikawaGenerator(api_key="sk-...", language="es")
>>> 
>>> # Generar análisis
>>> analysis = generator.generate_ishikawa("Defectos en el producto final")
>>> 
>>> # Exportar a Excel
>>> generator.export_to_excel(analysis, "analisis_ishikawa.xlsx")
>>> 
>>> # Exportar a Drawing.io
>>> generator.export_to_drawio(analysis, "diagrama_ishikawa.drawio")
"""
import pandas as pd
import json
import os
import base64
from typing import Dict, List, Optional, Callable
from openai import OpenAI


class IshikawaGenerator:
    """
    Generador de análisis de causa raíz usando Ishikawa + 5 Porqués con IA.
    
    Esta clase proporciona métodos para generar análisis de causa raíz combinando
    el diagrama de Ishikawa (espina de pescado) con la técnica de los 5 Porqués
    para identificar causas profundas de problemas.
    """
    
    # Categorías estándar de Ishikawa (6M)
    CATEGORIAS = [
        "Mano de Obra",
        "Métodos",
        "Máquinas",
        "Materiales",
        "Medición",
        "Medio Ambiente"
    ]
    
    def __init__(self, api_key: str, model: str = "gpt-5.4", language: str = "es"):
        """
        Inicializar el generador de Ishikawa + 5 Porqués.
        
        Parameters
        ----------
        api_key : str
            API key de OpenAI para autenticación
        model : str, optional
            Modelo de OpenAI a utilizar (default: "gpt-5.4")
        language : str, optional
            Idioma de salida del análisis (default: "es")
            Opciones: 'es' (español), 'en' (inglés)
        """
        self.client = OpenAI(api_key=api_key)
        self.model = model
        self.language = language
        self.prompts_dir = os.path.join(os.path.dirname(__file__), 'prompts')
    
    def load_prompt_template(self, language: str) -> str:
        """
        Cargar plantilla de prompt desde archivo markdown.
        
        Parameters
        ----------
        language : str
            Código de idioma de la plantilla ('es' o 'en')
            
        Returns
        -------
        str
            Contenido de la plantilla de prompt
        """
        prompt_file = os.path.join(self.prompts_dir, f'ishikawa_prompt_{language}.md')
        
        try:
            with open(prompt_file, 'r', encoding='utf-8') as f:
                return f.read()
        except FileNotFoundError:
            raise FileNotFoundError(f"No se encontró el archivo de prompt: {prompt_file}")
        except Exception as e:
            raise Exception(f"Error al leer el archivo de prompt: {str(e)}")
    
    def _process_image_file(self, file_data: Dict) -> Dict:
        """
        Procesar archivo de imagen para envío a la API.
        
        Parameters
        ----------
        file_data : Dict
            Diccionario con 'name', 'type', y 'content' (bytes) del archivo
            
        Returns
        -------
        Dict
            Diccionario con información procesada de la imagen
        """
        # Codificar imagen en base64
        base64_image = base64.b64encode(file_data['content']).decode('utf-8')
        
        # Determinar el tipo MIME
        file_ext = file_data['name'].lower().split('.')[-1]
        mime_types = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'webp': 'image/webp'
        }
        mime_type = mime_types.get(file_ext, 'image/jpeg')
        
        return {
            'type': 'image',
            'name': file_data['name'],
            'data': f"data:{mime_type};base64,{base64_image}"
        }
    
    def _process_document_file(self, file_data: Dict) -> Dict:
        """
        Procesar archivo de documento (txt, pdf, docx) para envío a la API.
        
        Parameters
        ----------
        file_data : Dict
            Diccionario con 'name', 'type', y 'content' (bytes) del archivo
            
        Returns
        -------
        Dict
            Diccionario con información procesada del documento
        """
        file_ext = file_data['name'].lower().split('.')[-1]
        
        try:
            if file_ext == 'txt':
                # Archivo de texto plano
                text = file_data['content'].decode('utf-8', errors='ignore')
                return {
                    'type': 'text',
                    'name': file_data['name'],
                    'content': text
                }
            
            elif file_ext == 'pdf':
                # PDF - enviar como base64 directamente al LLM
                base64_pdf = base64.b64encode(file_data['content']).decode('utf-8')
                return {
                    'type': 'document',
                    'name': file_data['name'],
                    'data': f"data:application/pdf;base64,{base64_pdf}"
                }
            
            elif file_ext == 'docx':
                # Word document - enviar como base64 directamente al LLM
                base64_docx = base64.b64encode(file_data['content']).decode('utf-8')
                return {
                    'type': 'document',
                    'name': file_data['name'],
                    'data': f"data:application/vnd.openxmlformats-officedocument.wordprocessingml.document;base64,{base64_docx}"
                }
            
            else:
                return {
                    'type': 'text',
                    'name': file_data['name'],
                    'content': f"[Tipo de archivo no soportado: {file_ext}]"
                }
        
        except Exception as e:
            return {
                'type': 'text',
                'name': file_data['name'],
                'content': f"[Error al procesar archivo: {str(e)}]"
            }
    
    def generate_ishikawa(self, 
                         problema: str,
                         contexto: str = "",
                         num_causas_por_categoria: int = 3,
                         uploaded_files: Optional[List[Dict]] = None,
                         progress_callback: Optional[Callable] = None) -> Dict:
        """
        Generar análisis de Ishikawa + 5 Porqués para un problema.
        
        Parameters
        ----------
        problema : str
            Descripción del problema a analizar
        contexto : str, optional
            Contexto adicional sobre el problema o proceso
        num_causas_por_categoria : int, optional
            Número de causas a identificar por cada categoría (default: 3)
        uploaded_files : List[Dict], optional
            Lista de archivos subidos (imágenes, PDFs, documentos)
            Cada diccionario debe tener 'name', 'type', y 'content' (bytes)
        progress_callback : callable, optional
            Función para reportar progreso (recibe float 0-1 y str mensaje)
            
        Returns
        -------
        Dict
            Diccionario con la estructura del análisis completo
        """
        if progress_callback:
            progress_callback(0.1, "Cargando plantilla de análisis...")
        
        # Cargar plantilla de prompt
        prompt_template = self.load_prompt_template(self.language)
        
        # Procesar archivos subidos si existen
        additional_context = ""
        files_for_api = []
        
        if uploaded_files:
            if progress_callback:
                progress_callback(0.15, f"Procesando {len(uploaded_files)} archivo(s)...")
            
            for file_data in uploaded_files:
                file_ext = file_data['name'].lower().split('.')[-1]
                
                # Determinar si es imagen o documento
                if file_ext in ['jpg', 'jpeg', 'png', 'webp']:
                    # Procesar imagen
                    processed = self._process_image_file(file_data)
                    files_for_api.append(processed)
                    
                elif file_ext in ['txt', 'pdf', 'docx']:
                    # Procesar documento
                    processed = self._process_document_file(file_data)
                    
                    if processed['type'] == 'document':
                        # PDFs y DOCX se envían como archivos al API
                        files_for_api.append(processed)
                    else:
                        # Archivos de texto plano se agregan al contexto
                        additional_context += f"\n\n--- Contenido de '{processed['name']}' ---\n{processed['content']}\n"
        
        # Construir el prompt con los datos
        full_context = contexto if contexto else "No se proporcionó contexto adicional."
        if additional_context:
            full_context += additional_context
        
        prompt = prompt_template.format(
            problema=problema,
            contexto=full_context,
            num_causas=num_causas_por_categoria
        )
        
        if progress_callback:
            progress_callback(0.3, "Generando análisis de causa raíz con IA...")
        
        try:
            # Preparar mensajes para la API
            messages = [
                {
                    "role": "system",
                    "content": "Eres un experto en análisis de causa raíz usando metodología Ishikawa y 5 Porqués. Generas análisis detallados y estructurados en formato JSON."
                }
            ]
            
            # Si hay archivos (imágenes o documentos), usar formato de contenido multimodal
            if files_for_api:
                user_content = [
                    {
                        "type": "text",
                        "text": prompt
                    }
                ]
                
                # Agregar archivos (imágenes, PDFs, documentos)
                for file_item in files_for_api:
                    if file_item['type'] == 'image':
                        user_content.append({
                            "type": "image_url",
                            "image_url": {
                                "url": file_item['data']
                            }
                        })
                    elif file_item['type'] == 'document':
                        user_content.append({
                            "type": "image_url",
                            "image_url": {
                                "url": file_item['data']
                            }
                        })
                
                messages.append({
                    "role": "user",
                    "content": user_content
                })
            else:
                # Solo texto
                messages.append({
                    "role": "user",
                    "content": prompt
                })
            
            # Llamar a la API de OpenAI
            response = self.client.chat.completions.create(
                model=self.model,
                messages=messages,
                response_format={"type": "json_object"},
                temperature=0.7
            )
            
            if progress_callback:
                progress_callback(0.7, "Procesando resultados...")
            
            # Extraer respuesta
            result_text = response.choices[0].message.content
            result = json.loads(result_text)
            
            if progress_callback:
                progress_callback(1.0, "¡Análisis completado exitosamente!")
            
            return result
            
        except json.JSONDecodeError as e:
            raise ValueError(f"Error al decodificar respuesta JSON: {str(e)}")
        except Exception as e:
            raise Exception(f"Error al generar análisis: {str(e)}")
    
    def export_to_excel(self, analysis: Dict, output_file: str) -> str:
        """
        Exportar análisis a archivo Excel estructurado.
        
        Crea un archivo Excel con múltiples hojas:
        - Resumen: problema y resumen ejecutivo
        - Ishikawa: tabla con todas las causas por categoría
        - 5 Porqués: análisis detallado con todos los niveles
        
        Parameters
        ----------
        analysis : Dict
            Diccionario con el análisis generado
        output_file : str
            Ruta del archivo Excel de salida
            
        Returns
        -------
        str
            Ruta del archivo generado
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # --- Hoja 1: Resumen ---
        ws_resumen = wb.create_sheet("Resumen")
        ws_resumen.column_dimensions['A'].width = 20
        ws_resumen.column_dimensions['B'].width = 80
        
        # Título
        ws_resumen['A1'] = "Análisis de Causa Raíz"
        ws_resumen['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_resumen['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_resumen.merge_cells('A1:B1')
        
        # Información del problema
        row = 3
        ws_resumen[f'A{row}'] = "Problema:"
        ws_resumen[f'A{row}'].font = Font(bold=True)
        ws_resumen[f'B{row}'] = analysis.get('problema', '')
        ws_resumen[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        row += 2
        ws_resumen[f'A{row}'] = "Metodología:"
        ws_resumen[f'A{row}'].font = Font(bold=True)
        ws_resumen[f'B{row}'] = "Diagrama de Ishikawa (6M) + 5 Porqués"
        
        row += 2
        ws_resumen[f'A{row}'] = "Fecha de análisis:"
        ws_resumen[f'A{row}'].font = Font(bold=True)
        from datetime import datetime
        ws_resumen[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # --- Hoja 2: Diagrama Ishikawa (Resumen) ---
        ws_ishikawa = wb.create_sheet("Diagrama Ishikawa")
        
        # Encabezados
        headers = ["Categoría", "Causa Principal", "Descripción"]
        for col, header in enumerate(headers, 1):
            cell = ws_ishikawa.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Anchos de columna
        ws_ishikawa.column_dimensions['A'].width = 20
        ws_ishikawa.column_dimensions['B'].width = 40
        ws_ishikawa.column_dimensions['C'].width = 60
        
        # Llenar datos
        row = 2
        for categoria_data in analysis.get('categorias', []):
            categoria = categoria_data.get('categoria', '')
            for causa in categoria_data.get('causas', []):
                ws_ishikawa.cell(row, 1, categoria)
                ws_ishikawa.cell(row, 2, causa.get('causa_principal', ''))
                ws_ishikawa.cell(row, 3, causa.get('descripcion', ''))
                
                # Estilo
                for col in range(1, 4):
                    ws_ishikawa.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')
                    ws_ishikawa.cell(row, col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                row += 1
        
        # --- Hoja 3: 5 Porqués Detallado ---
        ws_porques = wb.create_sheet("5 Porqués")
        
        # Encabezados
        headers_porques = ["Categoría", "Causa Principal", "Por qué 1", "Por qué 2", "Por qué 3", "Por qué 4", "Por qué 5", "Causa Raíz"]
        for col, header in enumerate(headers_porques, 1):
            cell = ws_porques.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Anchos de columna
        for i in range(1, 9):
            ws_porques.column_dimensions[get_column_letter(i)].width = 25
        
        # Llenar datos
        row = 2
        for categoria_data in analysis.get('categorias', []):
            categoria = categoria_data.get('categoria', '')
            for causa in categoria_data.get('causas', []):
                ws_porques.cell(row, 1, categoria)
                ws_porques.cell(row, 2, causa.get('causa_principal', ''))
                
                porques = causa.get('cinco_porques', [])
                for i, porque in enumerate(porques, 3):
                    ws_porques.cell(row, i, porque)
                
                ws_porques.cell(row, 8, causa.get('causa_raiz', ''))
                
                # Estilo
                for col in range(1, 9):
                    ws_porques.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')
                    ws_porques.cell(row, col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Resaltar causa raíz
                ws_porques.cell(row, 8).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                ws_porques.cell(row, 8).font = Font(bold=True)
                
                row += 1
        
        # Guardar archivo
        wb.save(output_file)
        return output_file
    
    def export_to_drawio(self, analysis: Dict, output_file: str) -> str:
        """
        Exportar análisis a formato Drawing.io (diagrama.drawio).
        
        Crea un diagrama visual de Ishikawa (espina de pescado) con todas
        las causas organizadas por categoría.
        
        Parameters
        ----------
        analysis : Dict
            Diccionario con el análisis generado
        output_file : str
            Ruta del archivo .drawio de salida
            
        Returns
        -------
        str
            Ruta del archivo generado
        """
        import xml.etree.ElementTree as ET
        from xml.dom import minidom
        import urllib.parse
        
        # Crear estructura XML básica de Draw.io
        mxfile = ET.Element('mxfile')
        mxfile.set('host', 'app.diagrams.net')
        mxfile.set('modified', '2024-01-01T00:00:00.000Z')
        mxfile.set('agent', 'Python IshikawaGenerator')
        mxfile.set('version', '22.0.0')
        mxfile.set('type', 'device')
        
        diagram = ET.SubElement(mxfile, 'diagram')
        diagram.set('id', 'ishikawa-diagram')
        diagram.set('name', 'Diagrama Ishikawa')
        
        mxGraphModel = ET.SubElement(diagram, 'mxGraphModel')
        mxGraphModel.set('dx', '1422')
        mxGraphModel.set('dy', '794')
        mxGraphModel.set('grid', '1')
        mxGraphModel.set('gridSize', '10')
        mxGraphModel.set('guides', '1')
        mxGraphModel.set('tooltips', '1')
        mxGraphModel.set('connect', '1')
        mxGraphModel.set('arrows', '1')
        mxGraphModel.set('fold', '1')
        mxGraphModel.set('page', '1')
        mxGraphModel.set('pageScale', '1')
        mxGraphModel.set('pageWidth', '1600')
        mxGraphModel.set('pageHeight', '900')
        
        root = ET.SubElement(mxGraphModel, 'root')
        
        # Celdas base
        cell0 = ET.SubElement(root, 'mxCell')
        cell0.set('id', '0')
        
        cell1 = ET.SubElement(root, 'mxCell')
        cell1.set('id', '1')
        cell1.set('parent', '0')
        
        cell_id = 2
        
        # Configuración de posiciones
        problema_x = 1200
        problema_y = 400
        spine_start_x = 200
        spine_end_x = problema_x
        spine_y = problema_y
        
        # Dibujar línea principal (espina central)
        spine = ET.SubElement(root, 'mxCell')
        spine.set('id', str(cell_id))
        spine.set('value', '')
        spine.set('style', 'endArrow=classic;html=1;strokeWidth=3;strokeColor=#000000;')
        spine.set('edge', '1')
        spine.set('parent', '1')
        
        geometry = ET.SubElement(spine, 'mxGeometry')
        geometry.set('width', '50')
        geometry.set('height', '50')
        geometry.set('relative', '1')
        geometry.set('as', 'geometry')
        
        source_point = ET.SubElement(geometry, 'mxPoint')
        source_point.set('x', str(spine_start_x))
        source_point.set('y', str(spine_y))
        source_point.set('as', 'sourcePoint')
        
        target_point = ET.SubElement(geometry, 'mxPoint')
        target_point.set('x', str(spine_end_x))
        target_point.set('y', str(spine_y))
        target_point.set('as', 'targetPoint')
        
        cell_id += 1
        
        # Dibujar cabeza (problema)
        problema_text = analysis.get('problema', 'Problema')
        problema_cell = ET.SubElement(root, 'mxCell')
        problema_cell.set('id', str(cell_id))
        problema_cell.set('value', problema_text)
        problema_cell.set('style', 'ellipse;whiteSpace=wrap;html=1;fillColor=#FFE6CC;strokeColor=#D79B00;strokeWidth=2;fontSize=14;fontStyle=1;')
        problema_cell.set('vertex', '1')
        problema_cell.set('parent', '1')
        
        geometry = ET.SubElement(problema_cell, 'mxGeometry')
        geometry.set('x', str(problema_x))
        geometry.set('y', str(problema_y - 60))
        geometry.set('width', '200')
        geometry.set('height', '120')
        geometry.set('as', 'geometry')
        
        cell_id += 1
        
        # Dibujar categorías y causas
        categorias = analysis.get('categorias', [])
        num_categorias = len(categorias)
        
        # Dividir categorías: mitad arriba, mitad abajo
        categorias_arriba = categorias[:num_categorias//2]
        categorias_abajo = categorias[num_categorias//2:]
        
        # Posiciones para categorías
        spacing_x = (spine_end_x - spine_start_x) / (max(len(categorias_arriba), len(categorias_abajo)) + 1)
        
        # Función para dibujar categoría y sus causas
        def draw_categoria(categoria_data, x_pos, y_offset, is_top):
            nonlocal cell_id
            
            categoria_nombre = categoria_data.get('categoria', '')
            causas = categoria_data.get('causas', [])
            
            # Punto en la espina principal
            spine_point_y = spine_y
            
            # Línea de categoría (diagonal)
            angle_y = spine_point_y + y_offset
            
            cat_line = ET.SubElement(root, 'mxCell')
            cat_line.set('id', str(cell_id))
            cat_line.set('value', '')
            cat_line.set('style', 'endArrow=classic;html=1;strokeWidth=2;strokeColor=#000000;')
            cat_line.set('edge', '1')
            cat_line.set('parent', '1')
            
            geometry = ET.SubElement(cat_line, 'mxGeometry')
            geometry.set('width', '50')
            geometry.set('height', '50')
            geometry.set('relative', '1')
            geometry.set('as', 'geometry')
            
            source_point = ET.SubElement(geometry, 'mxPoint')
            source_point.set('x', str(int(x_pos)))
            source_point.set('y', str(spine_point_y))
            source_point.set('as', 'sourcePoint')
            
            target_point = ET.SubElement(geometry, 'mxPoint')
            target_point.set('x', str(int(x_pos)))
            target_point.set('y', str(int(angle_y)))
            target_point.set('as', 'targetPoint')
            
            cell_id += 1
            
            # Etiqueta de categoría
            cat_label = ET.SubElement(root, 'mxCell')
            cat_label.set('id', str(cell_id))
            cat_label.set('value', categoria_nombre)
            cat_label.set('style', 'rounded=1;whiteSpace=wrap;html=1;fillColor=#DAE8FC;strokeColor=#6C8EBF;strokeWidth=2;fontSize=12;fontStyle=1;')
            cat_label.set('vertex', '1')
            cat_label.set('parent', '1')
            
            geometry = ET.SubElement(cat_label, 'mxGeometry')
            label_y = angle_y + (20 if not is_top else -80)
            geometry.set('x', str(int(x_pos - 60)))
            geometry.set('y', str(int(label_y)))
            geometry.set('width', '120')
            geometry.set('height', '40')
            geometry.set('as', 'geometry')
            
            cell_id += 1
            
            # Dibujar causas
            causa_y_start = label_y + (60 if not is_top else -120)
            for i, causa in enumerate(causas[:3]):  # Máximo 3 causas por categoría para claridad visual
                causa_text = causa.get('causa_principal', '')[:50] + '...' if len(causa.get('causa_principal', '')) > 50 else causa.get('causa_principal', '')
                
                causa_cell = ET.SubElement(root, 'mxCell')
                causa_cell.set('id', str(cell_id))
                causa_cell.set('value', causa_text)
                causa_cell.set('style', 'rounded=0;whiteSpace=wrap;html=1;fillColor=#FFF2CC;strokeColor=#D6B656;fontSize=10;')
                causa_cell.set('vertex', '1')
                causa_cell.set('parent', '1')
                
                geometry = ET.SubElement(causa_cell, 'mxGeometry')
                causa_y = causa_y_start + (i * 60 if not is_top else -(i * 60))
                geometry.set('x', str(int(x_pos - 80)))
                geometry.set('y', str(int(causa_y)))
                geometry.set('width', '160')
                geometry.set('height', '50')
                geometry.set('as', 'geometry')
                
                cell_id += 1
                
                # Línea conectando causa con categoría
                causa_line = ET.SubElement(root, 'mxCell')
                causa_line.set('id', str(cell_id))
                causa_line.set('value', '')
                causa_line.set('style', 'endArrow=classic;html=1;strokeWidth=1;strokeColor=#666666;dashed=1;')
                causa_line.set('edge', '1')
                causa_line.set('parent', '1')
                
                geometry = ET.SubElement(causa_line, 'mxGeometry')
                geometry.set('width', '50')
                geometry.set('height', '50')
                geometry.set('relative', '1')
                geometry.set('as', 'geometry')
                
                source_point = ET.SubElement(geometry, 'mxPoint')
                source_point.set('x', str(int(x_pos)))
                source_point.set('y', str(int(causa_y + 25)))
                source_point.set('as', 'sourcePoint')
                
                target_point = ET.SubElement(geometry, 'mxPoint')
                target_point.set('x', str(int(x_pos)))
                target_point.set('y', str(int(angle_y)))
                target_point.set('as', 'targetPoint')
                
                cell_id += 1
        
        # Dibujar categorías superiores
        for i, cat in enumerate(categorias_arriba):
            x_pos = spine_start_x + spacing_x * (i + 1)
            draw_categoria(cat, x_pos, -150, True)
        
        # Dibujar categorías inferiores
        for i, cat in enumerate(categorias_abajo):
            x_pos = spine_start_x + spacing_x * (i + 1)
            draw_categoria(cat, x_pos, 150, False)
        
        # Convertir a string XML formateado
        xml_string = ET.tostring(mxfile, encoding='unicode')
        
        # Formatear XML
        dom = minidom.parseString(xml_string)
        pretty_xml = dom.toprettyxml(indent='  ')
        
        # Guardar archivo
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)
        
        return output_file
